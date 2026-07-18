from openpyxl.styles import PatternFill, Font, Border, Side, Alignment
from openpyxl.utils import get_column_letter
from openpyxl import Workbook, load_workbook
from datetime import datetime
from os import path as os_path

class Color_style():
  PALE_MINT = PatternFill(fill_type="solid", fgColor="EBF1DE") 
  PALE_PINK = PatternFill(fill_type="solid", fgColor="F2DCDB")
  LIGHT_BLUE = PatternFill(fill_type="solid",fgColor="DCE6F1")
  SKY_BLUE = PatternFill(fill_type="solid",fgColor="C5D9F1")
  PEACH = PatternFill(fill_type="solid", fgColor="FDE9D9") 
  LACENDER = PatternFill(fill_type="solid", fgColor="E4DFEC")
  BEIGE = PatternFill(fill_type="solid",fgColor="DDD9C4")
  LIGHT_CORAL = PatternFill(fill_type="solid",fgColor="FFBDBD")
  YELLOW_GREEN = PatternFill(fill_type="solid",fgColor="92D050")
  SATIN_GREEN = PatternFill(fill_type="solid", fgColor="C6E0B4") 
  SPEARMINT_GREEN = PatternFill(fill_type="solid", fgColor="19ff78")
  FLIRTY_SALMON = PatternFill(fill_type="solid", fgColor="ff6d6d")

class Font_style():
  FONT = Font(name='Book Antiqua', size=12)

class Border_style():
  S_M = Side(border_style="medium")
  S_D = Side(border_style="dashed")
  S_T = Side(border_style="thin")

  border_dict = {
    "M" : S_M,
    "D" : S_D,
    "T" : S_T
  }

  @classmethod
  def set(cls,t=None,r=None,l=None,b=None): 
    return Border(
      top = cls.border_dict.get(t),         
      bottom = cls.border_dict.get(b),         
      left = cls.border_dict.get(l),         
      right = cls.border_dict.get(r)         
    )

class Function():
  @classmethod
  def init_freeze_row(self, ws, row_num):
    ws.freeze_panes = f"A{row_num+1}"


  @classmethod
  def read_first_line_of_file(self, file):
    with open(file, 'r', encoding='utf-8') as f:
      return f.readline().strip()

  @classmethod
  def is_date(self, data):
    try:
      datetime.strptime(data, "%d/%m/%Y")
      return True
    except ValueError:
      return False

  @classmethod
  def check_and_open_excel(self, file):
    if os_path.exists(file):
      wb = load_workbook(file)
    else:
      wb = Workbook()
    return wb

  @classmethod
  def read_file_data(self, file):
    with open(file,'r', encoding="utf-8") as file:
      data = file.read().splitlines()
    return data

  @classmethod
  def month_int_to_string(self, month):
    month_dict = {
      1: "Jan", 2: "Feb", 3: "Mar", 4: "Apr", 5: "May", 6: "Jun",
      7: "Jul", 8: "Aug", 9: "Sep", 10: "Oct", 11: "Nov", 12: "Dec"
    }
    return month_dict.get(month)

  @classmethod
  def open_month_ws(self, wb, ws):
    if ws in wb.sheetnames:
      return wb[ws]
    else:
      return wb.create_sheet(ws)

  @classmethod
  def get_max_row_by_value(self ,ws):
    for row in range(ws.max_row, 0, -1):
      if any(cell.value is not None for cell in ws[row]):
        return row
    return 0

  @classmethod
  def is_first_time_of_ws(self, ws, pass_number):
    last_row = self.get_max_row_by_value(ws)
    for i in range(last_row,1,-1):
      cell = f"B{i}"
      value = ws[cell].value
      if value != None:
        if self.is_date(ws[cell].value):
          if(pass_number == 0):
            return False
          else:
            pass_number-=1
    return True
  
  @classmethod
  def cell_type_message(self, ws,item): # item = [ [row(int), col(int), value(String)],[row1, col1, value1],...   ]
    for col, row, message in item:
      cell = f"{get_column_letter(col)}{row}"
      ws[cell].value = message
      ws[cell].font = Font_style.FONT

  @classmethod
  def range_cell_set_color(self, ws, item): # item = [ [col1,row1,col2,row2,color1],[col1,row1,col2,row2,color]]
    for col1, row1, col2, row2, color in item:
      for row in ws.iter_rows(min_col = col1 , min_row = row1,
                             max_col = col2 , max_row = row2):
        for cell in row:
          cell.fill = color

  @classmethod
  def cell_set_border(self, ws, item): # item = [ [col1, row1, border1], [col2, row2, border2], ...]
    for col, row, border in item:
      cell = f"{get_column_letter(col)}{row}"
      ws[cell].border = border 

  @classmethod
  def range_merge_cell(self, ws, item): # item = [ [[col1, row1 (start cell)],[col2, row2 (end cell)]], ... ]
    for start , end in item:
      start_cell = f"{get_column_letter(start[0])}{start[1]}"
      end_cell = f"{get_column_letter(end[0])}{end[1]}"
      ws.merge_cells(f"{start_cell}:{end_cell}")
      ws[start_cell].alignment = Alignment(horizontal='center', vertical='top')
  
  @classmethod
  def is_date_duplicate(self, ws, date):
    last_row = self.get_max_row_by_value(ws)
    for row in range(last_row, 1,-1):
      cell = f"B{row}"
      value = ws[cell].value
      if value != None:
        if self.is_date(value):
          if (value == date):
            return True
          else:
            return False

  @classmethod
  def find_start_row(self, ws):
    last_row = self.get_max_row_by_value(ws)

    for i in range(last_row, 1, -1):
      if any(cell.value is not None for cell in ws[i]):
        return i+1
    return 5

  @classmethod
  def input_date(self, ws, message):
    start_row = self.find_start_row(ws)
    cell_value = [[2,start_row, message]]
    self.cell_type_message(ws, cell_value)
    return start_row

  @classmethod
  def is_int(self, value):
    try:
      return float(value).is_integer()
    except (ValueError, TypeError):
      return False

  @classmethod
  def find_last_value_row(self, ws, value, row, last_find_row):
    for i in range(row, last_find_row, -1):
      cell = f"B{i}" 
      cell_value = ws[cell].value
      if (cell_value == value):
        return i
    return None

  @classmethod
  def is_num(self, num):
    try:
      float(num)  
      return True
    except ValueError:
      return False


if __name__ == "__main__":
  pass