from openpyxl import load_workbook
from openpyxl.styles import Alignment
from openpyxl.utils import get_column_letter
from os import (path as os_path,
                makedirs as os_makedirs)
from decimal import Decimal
from shutil import (copy as shutil_copy)
from shlex import (split as shlex_split)

import config
from excel_helper import (Color_style as eh_color,
                          Font_style as eh_font,
                          Border_style as eh_border,
                          Function as eh_funct)

INPUT_FOLDER_PATH = config.path.INPUT_FOLDER
MONEY_MESSAGE_FILE_PATH = os_path.join(INPUT_FOLDER_PATH, config.path.EXCEL_INPUT_FILE_LIST[0])  
CURRENT_HAVE_FILE_PATH = os_path.join(INPUT_FOLDER_PATH, config.path.EXCEL_INPUT_FILE_LIST[1])

OUTPUT_FOLDER_PATH = config.path.OUTPUT_FOLDER

message_out = print
def set_message_out(function):
    global message_out
    message_out = function

def get_last_month_money_message(wb, month, year):
    if(month == "1" or month == "01"):
        try:
            last_year = int(year) - 1 
            last_year_wb = load_workbook(os_path.join(OUTPUT_FOLDER_PATH, str(last_year),"money","money.xlsx"))
            last_month_ws = last_year_wb["Dec"]

        except FileNotFoundError:
            return None
    else:
        try:
            last_month_ws = wb[eh_funct.month_int_to_string(int(month)-1)]
        except KeyError:
            return None

    ws_last_row = eh_funct.get_max_row_by_value(last_month_ws)
    cell = f"B{ws_last_row}"
    ws_last_row_colB_value = last_month_ws[cell].value


    if(ws_last_row_colB_value != "current have"):
        ws_last_row = eh_funct.find_last_value_row(last_month_ws, "sum", ws_last_row)

    result = []
    for col in [3,5,7,9,11,13,15,21]:
        cell = f"{get_column_letter(col)}{ws_last_row}"
        value = last_month_ws[cell].value
        result.append(value)
    return result 



def sheet_init(wb, ws, month, year): # funciton for the first day of every month
    ws.column_dimensions['B'].width = 14.5

    for i in range(3,23):
        col = get_column_letter(i)
        ws.column_dimensions[col].width = 12

    last_month_have = get_last_month_money_message(wb, month, year)

    cell_message = [ 
        [2, 2, "Date"], [3 , 2, "Cash"], [5,2, "Coin-hold"], [7, 2, "HSBC"], [9, 2, "Octopus"], [11, 2, "AliPay"], [13, 2, "Tap and Go"], [15, 2, "Payme"], [ 17, 2, "Countable Total"], [19, 2, "Real Total"], [22, 2, "Correct"], 
        [2, 3, "DD/MM/YY"], [3, 3, 'I'], [4, 3, 'O'], [5, 3, 'I'], [6, 3, 'O'], [7, 3, 'I'], [8, 3, 'O'], [9, 3, 'I'], [10, 3, 'O'], [11, 3, 'I'], [12, 3, 'O'], [13, 3, 'I'], [14, 3, 'O'], [15, 3, 'I'], [16, 3, 'O'], [17, 3, 'I'], [18, 3, 'O'], [19, 3, 'I'], [20, 3, 'O'], [21, 3, "Have"]
    ]

    if(last_month_have != None):
        last_message =[2, 4, "Last"], [3, 4, last_month_have[0]], [5, 4,last_month_have[1]], [7, 4,last_month_have[2] ], [9, 4,last_month_have[3] ], [11, 4,last_month_have[4] ], [13, 4,last_month_have[5] ],[15,4,last_month_have[6]],[ 21, 4,last_month_have[7] ]
    else:
        last_message= [2, 4, "Last"], [3, 4, "None"], [5, 4, "None"], [7, 4, "None"], [9, 4, "None"], [11, 4, "None"], [13, 4, "None"],[15, 4, "None"], [ 21, 4, "None"]
    
    cell_message.extend(last_message)

    cell_color = [
        [3, 2, 4, 4, eh_color.PALE_MINT],
        [5, 2, 6, 4, eh_color.SATIN_GREEN],
        [7, 2, 8, 4, eh_color.PALE_PINK],
        [9, 2, 10, 4, eh_color.LIGHT_BLUE],
        [11, 2, 12, 4, eh_color.SKY_BLUE],
        [13, 2, 14, 4, eh_color.PEACH],
        [15, 2, 16, 4, eh_color.LACENDER],
        [17, 2, 18, 4, eh_color.BEIGE],
        [19, 2, 21, 4, eh_color.LIGHT_CORAL],
        [22, 2, 22, 4, eh_color.YELLOW_GREEN],
    ]

    cell_border = [ 
        *([2+i, 2 , eh_border.set("M","M","M","M") ] for i in range(0,21)),
        *([2+i, 2+j , eh_border.set("M","M","M","M")] for i in [0,20] for j in [1,2]),
        *([2+i, 2+j , eh_border.set("M","D","M","M")] for i in range(1,18) if i%2==1 for j in [1,2]),
        *([2+i, 2+j , eh_border.set("M","M","D","M")] for i in range(1,18) if i%2==0 for j in [1,2]),
        *([20, 2+j, eh_border.set("M","D","D","M")] for j in [1,2]),
        *([21, 2+j, eh_border.set("M","M","D","M")] for j in [1,2]),
    ] 

    merge_cell = [
        [[3, 2], [4, 2]],
        [[5, 2], [6, 2]],
        [[7, 2], [8, 2]],
        [[9, 2], [10, 2]],
        [[11, 2], [12, 2]],
        [[13, 2], [14, 2]],
        [[15, 2], [16, 2]],
        [[17, 2], [18, 2]],
        [[19, 2], [21, 2]]
    ]

    eh_funct.cell_type_message(ws,cell_message)
    eh_funct.range_cell_set_color(ws, cell_color)
    eh_funct.cell_set_border(ws, cell_border)
    eh_funct.range_merge_cell(ws, merge_cell)

ACCOUNT_LIST = [ ["C", 1], ["CH", 3] ,["H", 5], ["O", 7], ["AP", 9], ["TnG", 11], ["P", 13]]

def self_convert(ws, amount, account, to, record_list):
    account_add_amount = next(num for key,num in ACCOUNT_LIST if key == account)
    to_add_amount = next(num for key,num in ACCOUNT_LIST if key == to)
    col, row  = record_list[0], record_list[1]
    message = [
        [col+account_add_amount, row, to],
        [col+account_add_amount+1, row, amount],
        [col+to_add_amount, row, amount],
        [col+to_add_amount+1, row, account]
    ]
    eh_funct.cell_type_message(ws, message)

def income(ws, amount, account, to, countable, record_list):
    to_add_amount = next(num for key, num in ACCOUNT_LIST if key == to)  
    col , row = record_list[0],record_list[1]
    message = [
        [col+to_add_amount, row, amount],
        [col+to_add_amount+1, row, account],
        [col+15, row , amount],
        [col+16, row, account],
        [col+17, row, amount],
        [col+18, row, account]
    ]
    if(not countable):
        message.pop(2)
        message.pop(2) 
   
    eh_funct.cell_type_message(ws, message)

def pay(ws, amount, account, to, countable, record_list):
    account_add_amount = next(num for key,num in ACCOUNT_LIST if key == account)
    col, row = record_list[0], record_list[1]
    message = [
        [col+account_add_amount, row, to],
        [col+account_add_amount+1, row, amount],
        [col+15, row, to],
        [col+16, row, amount],
        [col+17, row, to],
        [col+18, row, amount]
    ]     
    if(not countable):
        message.pop(2)
        message.pop(2)

    eh_funct.cell_type_message(ws, message)

def cal_sum_in_outcome(ws, col, start_row, end_row):
    return_value = Decimal(0)
    for row in range(start_row, end_row):

        cell = f"{get_column_letter(col)}{row}"
        try:
            cell_value = ws[cell].value
            if(eh_funct.is_num(cell_value)):
                cal_value =  Decimal(cell_value)
            else:
                cal_value = Decimal('0')

        except TypeError:
            cal_value =  Decimal('0')
        return_value += cal_value
    return return_value

def day_summary(ws, date, end_row):
    date_start_row = eh_funct.find_last_value_row(ws, date, end_row)

    if (date_start_row == None):
        message_out(f"Money Excel: Error, cannot find last cell of value: {date}", color="red")

    if(eh_funct.is_first_time_of_ws(ws, 1)): # 1 mean only have one date, which is the current date, so if just only current date need to sum, this is the first time
        last_row = eh_funct.find_last_value_row(ws, "Last", end_row)
    else:
        last_row = eh_funct.find_last_value_row(ws, "sum", end_row)

    message = [ [2,end_row, "sum"]]
    eh_funct.cell_type_message(ws, message)

    total_money = Decimal('0')
    for pair in ACCOUNT_LIST:
        income_money = Decimal('0')
        outcome_money = Decimal('0')

        col = 2 + pair[1]

        last_cell = f"{get_column_letter(col)}{last_row}"
        last_money_raw_value = ws[last_cell].value 
        
        if (last_money_raw_value is None or last_money_raw_value == "None"):
            last_money = Decimal('0')
        else:
            last_money = Decimal(str(last_money_raw_value))

        income_money = cal_sum_in_outcome(ws, col, date_start_row, end_row) 
        outcome_money = cal_sum_in_outcome(ws, col+1, date_start_row, end_row)

        total_value = last_money +  income_money - outcome_money 
        total_money += total_value

        message = [[col, end_row, total_value]]

        eh_funct.cell_type_message(ws,message)

    message = [[21, end_row, total_money]]
    eh_funct.cell_type_message(ws, message)

    for col in [17,19]:
        total_input = cal_sum_in_outcome(ws, col , date_start_row, end_row)
        total_output = cal_sum_in_outcome(ws, col+1 , date_start_row, end_row)

        message = [[col, end_row, total_input],[col+1, end_row, total_output]]

        eh_funct.cell_type_message(ws, message)
    
    if(message_out):
        message_out("------------------------------------------------------------------")
        message_out(f"Excel money {date} input is \"{total_input}\"", new_line = True)
        message_out(f"Excel money {date} output is \"{total_output}\"", new_line = True)
    else:
        message_out(f"Excel money {date} input is \"{total_input}\"")
        message_out(f"Excel money {date} output is \"{total_output}\"")

    last_have_value = Decimal(str(ws[f"{get_column_letter(21)}{last_row}"].value)).quantize(Decimal("0.00"))
    current_input_value = Decimal(str(ws[f"{get_column_letter(19)}{end_row}"].value)).quantize(Decimal("0.00"))
    current_output_value = Decimal(str(ws[f"{get_column_letter(20)}{end_row}"].value)).quantize(Decimal("0.00"))
    current_have_value = Decimal(str(ws[f"{get_column_letter(21)}{end_row}"].value)).quantize(Decimal("0.00"))
    
    correct_value_bool = last_have_value == current_have_value + current_output_value - current_input_value
    if (not correct_value_bool):
        message_out("Money excel summary have false", color="red")
        return False

    correct_value = str(correct_value_bool)
    message = [[22, end_row, correct_value]]
    eh_funct.cell_type_message(ws, message)
    return True

def day_total(ws, row, summary_correct):
    current_real_sum_in_cell = f"S{row-1}"
    current_real_sum_out_cell = f"T{row-1}"
    current_real_sum_in_value = Decimal(str(ws[current_real_sum_in_cell].value))
    current_real_sum_out_value = Decimal(str(ws[current_real_sum_out_cell].value))

    current_countable_sum_in_cell = f"Q{row-1}"
    current_countable_sum_out_cell = f"R{row-1}"
    current_countable_sum_in_value = Decimal(str(ws[current_countable_sum_in_cell].value))
    current_countable_sum_out_value = Decimal(str(ws[current_countable_sum_out_cell].value))

    if(eh_funct.is_first_time_of_ws(ws, 1)):
        message = [[2, row, "total"], [17, row, current_countable_sum_in_value], [18, row, current_countable_sum_out_value],[19, row, current_real_sum_in_value], [20, row, current_real_sum_out_value]] 
    else:
        last_row = eh_funct.find_last_value_row(ws, "total", row)
        last_real_total_in_cell = f"S{last_row}" 
        last_real_total_out_cell = f"T{last_row}"
        last_real_total_in_value = Decimal(str(ws[last_real_total_in_cell].value))
        last_real_total_out_value = Decimal(str(ws[last_real_total_out_cell].value))

        last_countable_sum_in_cell = f"Q{last_row}"
        last_countable_sum_out_cell = f"R{last_row}"
        last_countable_sum_in_value = Decimal(str(ws[last_countable_sum_in_cell].value))
        last_countable_sum_out_value = Decimal(str(ws[last_countable_sum_out_cell].value))

        total_in_value = last_real_total_in_value + current_real_sum_in_value
        total_out_value = last_real_total_out_value + current_real_sum_out_value

        countable_total_in_value = last_countable_sum_in_value + current_countable_sum_in_value
        countable_total_out_value = last_countable_sum_out_value + current_countable_sum_out_value

        if(summary_correct):
            message_out(f"so far total income is \"{total_in_value}\"")
            if(message_out):
                message_out(f"and total output is \"{total_out_value}\"", new_line = True )
            else:
                message_out(f"and total output is \"{total_out_value}\"")

            message_out(f"and countable income is \"{countable_total_in_value}\"")

            if(message_out):
                message_out(f"countable outcome is \"{countable_total_out_value}\"", new_line = True)
            else:
                message_out(f"countable outcome is \"{countable_total_out_value}\"")

        message = [[2, row, "total"], [17, row, countable_total_in_value], [18, row, countable_total_out_value], [19, row, total_in_value], [20, row, total_out_value]]

    eh_funct.cell_type_message(ws, message)

def mark_current_amount(ws, date, row, summary_correct):
    if(os_path.exists(CURRENT_HAVE_FILE_PATH)):
        full_file = eh_funct.read_file_data(CURRENT_HAVE_FILE_PATH)

        message = [[2, row, "current have"]]
        is_current_day = False
        total_amount = Decimal("0.00")
        for line in full_file:
            line = line.strip()
            if( is_current_day == False and line == date):
                is_current_day = True
                continue
            elif(is_current_day):
                account , amount = line.split(' ')
                amount = Decimal(amount)
                col = next(num for key, num in ACCOUNT_LIST if account == key)

                sum_cell = f"{get_column_letter(2+col)}{row-2}"
                sum_value = Decimal(str(ws[sum_cell].value)).quantize(Decimal("0.00"))

                total_amount += amount

                correct_value_bool = (sum_value == amount)
                
                if(not correct_value_bool):
                    message_out(f"Money excel mark currnet amount for \"{account}\" in \"{date}\" have False", color="red")

                correct_value = "True" if (correct_value_bool) else "False"

                message.append([2+col, row, amount])
                message.append([2+col+1, row, correct_value]) 
            
            else:
                return False

        sum_row = eh_funct.find_last_value_row(ws, "sum", row)
        total_have_value = ws[f"U{sum_row}"].value

        is_equal = "True" if (total_have_value == total_amount) else "False"
        if(is_equal == "False"):
            message_out(f"Money excel current correct is False, total have value is {total_have_value} and total amount is {total_amount}", color="red")

        if(is_equal == "True" and summary_correct):
            message_out(f"total have money \"{total_have_value}\"")

        if(message_out):
                message_out("------------------------------------------------------------------")

        message.append([21, row, total_amount]) 
        message.append([22, row, is_equal])
        eh_funct.cell_type_message(ws, message)
        return True

    else:
        message_out("Money excel money_current_have.txt not exist", color="red")
        return False

def day_set_style(ws, date, row, current_marked):
    date_row = eh_funct.find_last_value_row(ws, date, row)
    #row = sum row, row + 1 = total row , row + 2 = current have row
    
    cell_border = [
        # the date_row style part
        *([i, date_row, eh_border.set("M","M","M","D")] for i in [2,22]),
        *([i, date_row, eh_border.set("M","D","M","D")] for i in range(3,20) if i%2==1),
        *([i, date_row, eh_border.set("M","M","D","D")] for i in range(4,21) if i%2==0),
        [20,date_row , eh_border.set("M","D","D","D")],
        # middle part 
        *([i, j, eh_border.set("D","M","M","D")] for i in [2,22] for j in range(date_row+1,row-1)), # only for the 2 and 20 col of the middle part
        *([i ,j, eh_border.set("D","D","M","D")] for i in range(3,20) if i%2==1 for j in range(date_row+1, row-1)),
        *([i, j, eh_border.set("D","M","D","D")] for i in range(4,22) if (i%2==0 or i==21) for j in range(date_row+1, row-1)),
        *([20,j, eh_border.set("D","D","D","D")] for j in range(date_row+1, row-1)), #only col 18
        # the row about the sum row
        *([i,row-1, eh_border.set("D","M","M","T")] for i in [2,22]),
        *([i, row-1, eh_border.set("D","D","M","T")] for i in range(3,20) if i%2==1),
        *([i, row-1, eh_border.set("D","M","D","T")] for i in range(4,22) if i%2==0 or i == 21),
        [20,row-1, eh_border.set("D","D","D","T")],
        # the sum row
        *([i,row, eh_border.set("T","M","M","T")] for i in [2,22]),
        *([i, row, eh_border.set("T","D","M","T")] for i in range(3,20) if i%2==1),
        *([i, row, eh_border.set("T","M","D","T")] for i in range(4,22) if i%2==0 or i == 21),
        [20,row, eh_border.set("T","D","D","T")],
        # the total row
        *([i,row+1, eh_border.set("T","M","M","M")] for i in [2,22]),
        *([i, row+1, eh_border.set("T","D","M","M")] for i in range(3,20) if i%2==1),
        *([i, row+1, eh_border.set("T","M","D","M")] for i in range(4,22) if i%2==0 or i == 21),
        [20,row+1, eh_border.set("T","D","D","M")]
    ]

    color_end_row = row+1
    if(current_marked):
        border_message = [
            *([i,row+2, eh_border.set("M","M","M","M")] for i in [2,22]),
            *([i, row+2, eh_border.set("M","D","M","M")] for i in range(3,20) if i%2==1),
            *([i, row+2, eh_border.set("M","M","D","M")] for i in range(4,22) if i%2==0 or i == 21),
            [20,row+2, eh_border.set("M","D","D","M")]
        ]
        cell_border.extend(border_message) 

        color_end_row = row+2

    cell_color= [
        [3, date_row, 4, color_end_row, eh_color.PALE_MINT],
        [5, date_row, 6, color_end_row, eh_color.SATIN_GREEN],
        [7, date_row, 8, color_end_row, eh_color.PALE_PINK],
        [9, date_row, 10, color_end_row, eh_color.LIGHT_BLUE],
        [11, date_row, 12, color_end_row, eh_color.SKY_BLUE],
        [13, date_row, 14, color_end_row, eh_color.PEACH],
        [15, date_row, 16, color_end_row, eh_color.LACENDER],
        [17, date_row, 18, color_end_row, eh_color.BEIGE],
        [19, date_row, 21, color_end_row, eh_color.LIGHT_CORAL],
        [22, date_row, 22, color_end_row, eh_color.YELLOW_GREEN]
    ]
    eh_funct.range_cell_set_color(ws, cell_color) 
    eh_funct.cell_set_border(ws,cell_border)

def day_finish(ws, date, row):
    summary_correct = day_summary(ws, date, row)
    day_total(ws, row+1, summary_correct)
    current_marked = mark_current_amount(ws, date, row+2, summary_correct)
    day_set_style(ws, date, row, current_marked)

def money_excel_process():
    message_out("Money excel processing")

    file_first_line = eh_funct.read_first_line_of_file(MONEY_MESSAGE_FILE_PATH)
    if((eh_funct.is_date(file_first_line))):
        f_day, f_month, f_year = file_first_line.split('/')
        MONEY_FOLDER_PATH = os_path.join(OUTPUT_FOLDER_PATH, f_year, "money")
        MONEY_FILE_PATH = os_path.join(MONEY_FOLDER_PATH, "money.xlsx")
        BACKUP_FILE_PATH = os_path.join(MONEY_FOLDER_PATH, "backup.xlsx")
    else:
        if(message_out):
            message_out("Money Excel : Error, the money_messages.txt first line is not a date!", color="red") 

    if(not os_path.exists(MONEY_FOLDER_PATH)):
        os_makedirs(MONEY_FOLDER_PATH)

    if(os_path.exists(MONEY_FILE_PATH)):
        shutil_copy(MONEY_FILE_PATH, BACKUP_FILE_PATH)

    wb = eh_funct.check_and_open_excel(MONEY_FILE_PATH)
    ws = wb.active

    current_date = None
    last_day = None
    money_message_list = eh_funct.read_file_data(MONEY_MESSAGE_FILE_PATH)
    START_COL = 2
    have_previous_date = False
    for message in money_message_list:
        if(message == ''):
            continue

        if eh_funct.is_date(message):

            if(current_date != None):
                last_day = current_date

            if(have_previous_date):
                day_finish(ws, last_day, start_row + record)
                
            day, month, year = message.split('/') 
            month_str = eh_funct.month_int_to_string(int(month))
            ws = eh_funct.open_month_ws(wb, month_str)

            current_date = message

            if(eh_funct.is_first_time_of_ws(ws,0)): 
                sheet_init(wb, ws, month, year)
            else:
                if(eh_funct.is_date_duplicate(ws,message)):
                    message_out(f"Money_excel : date {message} is marked!", color = "red")

            start_row = eh_funct.input_date(ws, message)
            have_previous_date = True
            record = 0
        else: 
            amount, account, to, countable, *rest = shlex_split(message)  # *rest is just let the code can run successfully 
            account = account.strip('"')
            account = account.strip()
            to = to.strip('"')
            to = to.strip()
            countable = (countable == "True")
            if any(to == pair[0] for pair in ACCOUNT_LIST) and any(account == pair[0] for pair in ACCOUNT_LIST ):
                self_convert(ws, amount, account, to, [START_COL, start_row + record])
            elif any(to == pair[0] for pair in ACCOUNT_LIST):
                income(ws,amount, account, to , countable, [START_COL, start_row + record])
            else:
                pay(ws, amount, account, to, countable, [START_COL, start_row + record])
            record+=1

    if(current_date is not None):
        day_finish(ws, current_date, start_row + record)

    wb.save(MONEY_FILE_PATH)
    message_out("Money excal process done")

    # --- stop copy --- #

if __name__ == "__main__":
    money_excel_process()