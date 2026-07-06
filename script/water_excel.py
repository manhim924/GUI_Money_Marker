from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment
from openpyxl.utils import get_column_letter
from sys import (exit as sys_exit)
from os import (path as os_path,
                makedirs as os_makedirs)
from datetime import datetime
from decimal import Decimal
from shutil import (copy as shutil_copy)
from shlex import (split as shlex_split)
import config
from excel_helper import (Color_style as eh_color,
                          Font_style as eh_font,
                          Border_style as eh_border,
                          Function as eh_funct)

INPUT_FOLDER_PATH = config.path.INPUT_FOLDER
WATER_MESSAGE_FILE_PATH = os_path.join(INPUT_FOLDER_PATH, config.path.EXCEL_INPUT_FILE_LIST[5])

OUTPUT_FOLDER_PATH = config.path.OUTPUT_FOLDER

CONTAINER_LIST = [ ["home", 950], ["outside", 1100]]

message_out = print
def set_message_out(function):
    global message_out
    message_out = function

def sheet_init(wb, ws, month, year):
    ws.column_dimensions['B'].width = 14.5

    for i in range(3,23):
        col = get_column_letter(i)
        ws.column_dimensions[col].width = 12

    cell_message = [
        [2, 2, "Date"], 
        [3, 2, "Water Fill"], 
        [4, 2, "ml"],
        [5, 2, "Aim"],
        [6, 2, "Total ml"]
    ]

    cell_color = [
        [3, 2, 3, 2, eh_color.PALE_MINT],
        [4, 2, 4, 2, eh_color.PALE_PINK],
        [5, 2, 5, 2, eh_color.LIGHT_BLUE],
        [6, 2, 6, 2, eh_color.SKY_BLUE]
    ]

    cell_border = [
        *([2+i , 2, eh_border.set("M", "M", "M", "M")] for i in range(0,5))
    ]

    eh_funct.cell_type_message(ws, cell_message)
    eh_funct.range_cell_set_color(ws, cell_color)
    eh_funct.cell_set_border(ws, cell_border)

def water_mark(ws, container, record_list):
    container_capacity = next(capacity_ for container_ , capacity_ in CONTAINER_LIST if container_ == container)
    col , row = record_list[0], record_list[1] 

    message = [
        [col+1, row, container],
        [col+2, row, container_capacity]
    ]

    eh_funct.cell_type_message(ws, message)



def get_day_total_ml(ws, date, end_row):
    return_value = Decimal(0)
    date_start_row = eh_funct.find_last_value_row(ws, date, end_row)

    if(date_start_row == None):
        if(not message_out):
            message_out(f"Water Excel: Error, cannot find last cell of value: {date}", color = "red")

    for row in range(date_start_row, end_row):
        cell = f"{get_column_letter("D")}{row}"

        try:
            cell_value = ws[cell].value

            if(eh_funct.is_num(cell_value)):
                cal_value = Decimal(cell_value)

            else:
                cal_value = Decimal('0')

        except TypeError:
            cal_value = Decimal('0')

        return_value += cal_value

    return return_value

def day_total(ws, date, row):
    target = 3000
    total_ml = get_day_total_ml(ws, date, row)
    
    target_aimed = True if target <= total_ml else False

    message = [
        [2, row, "total"],
        [4, row, total_ml],
        [5, row, target]
    ]

    if(not eh_funct.is_first_time_of_ws(ws, 1)):
        last_row = eh_funct.find_last_value_row(ws, "total", row)
        last_total_ml_cell = f"F{last_row}"
        last_total_ml_value = Decimal(str(ws[last_total_ml_cell].value))

        total_ml = total_ml+last_total_ml_value

    message.append([6, row, total_ml])
    
    eh_funct.cell_type_message(ws, message)

    return target_aimed

def day_set_style(ws, date, row, target_aimed):
    pass

def day_finish(ws, date, row):
    target_aimed = day_total(ws, date, row+1)    
    day_set_style(ws, date, row, target_aimed)

def water_excel_process():
    message_out("Water excel started!")

    file_first_line = eh_funct.read_first_line_of_file(WATER_MESSAGE_FILE_PATH)
    if(eh_funct.is_date(file_first_line)):
        f_day, f_month, f_year = file_first_line.split('/')
        WATER_FOLDER_PATH = os_path.join(OUTPUT_FOLDER_PATH, f_year, "water")
        WATER_FILE_PATH = os_path.join(WATER_FOLDER_PATH, "water.xlsx")
        BACKUP_FILE_PATH = os_path.join(WATER_FOLDER_PATH, "backup.xlsx")
    else:
        if(message_out):
            message_out("Water Excel : Error, the water_message.txt first line is not a date!")
            sys_exit(0)

    if(not os_path.exists(WATER_FOLDER_PATH)):
        os_makedirs(WATER_FOLDER_PATH)
    
    if(os_path.exists(WATER_FILE_PATH)):
        shutil_copy(WATER_FILE_PATH, BACKUP_FILE_PATH)

    wb = eh_funct.check_and_open_excel(WATER_FILE_PATH)
    ws = wb.active

    current_date = None
    last_day = None
    water_message_list = eh_funct.read_file_data(WATER_MESSAGE_FILE_PATH)
    START_COL = 2
    have_previous_date = False
    can_save = True

    for message in water_message_list: 
        if(message == ''):
            continue

        if eh_funct.is_date(message):
            if(current_date != None):
                last_day = current_date

            if(have_previous_date):
                day_finish(ws, last_day, start_row + record)

            day, month, year = message.split('/')
            month_str = eh_funct.month_int_to_string(int(month))
            ws = eh_funct.open_month_ws(wb, month_str)

            current_date = message

            if(eh_funct.is_first_time_of_ws(ws,0)):
                sheet_init(wb, ws, month, year)
                pass
            else:
                if(eh_funct.is_date_duplicate(ws, message)):
                    message_out(f"Water_excel : date {message} is marked", color = "red")

            start_row = eh_funct.input_date(ws, message)
            have_previous_date = True
            record = 0
        else:
            container, *rest = shlex_split(message) 
            container = container.strip('"')

            print(container)

            water_mark(ws, container, [START_COL, start_row + record])

            record+=1

    if(current_date is not None):
        day_finish(ws, current_date, start_row+ record)

    wb.save(WATER_FILE_PATH)
    message_out("Water excal process done")


if __name__ == "__main__":
    water_excel_process()